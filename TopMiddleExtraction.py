import openpyxl
from openpyxl import Workbook
from langchain_ollama import OllamaLLM

# Initialize the Phi4 model
model = OllamaLLM(model="Phi4", use_cuda=True)

# Load the input Excel file
input_file = "Data4Good_Arolsen_Archives_50k.xlsx"
output_file = "output.xlsx"

wb = openpyxl.load_workbook(input_file)
sheet = wb.active

# Prepare the output Excel workbook
output_wb = Workbook()
output_sheet = output_wb.active
output_sheet.title = "Extracted Data"

# Write headers to the output Excel file
headers = ["ID", "Name", "Surname", "Father", "Mother", "Spouse", "Birthplace", "Nationality", "Religion", "Post War Occupation"]
output_sheet.append(headers)

# Iterate through the rows of the input file, starting after the headers
for row in sheet.iter_rows(min_row=2, values_only=True):
    print(f"Processing row: {row}")  # Print the current row being processed
    id_value = row[0]  # Assuming 'ID' is the first column
    upper_text = row[15]  # Assuming 'Upper' is the second column
    middle_text = row[16]  # Assuming 'Middle' is the third column

    # Combine the 'Upper' and 'Middle' columns for context
    context = f"Upper: {upper_text}\nMiddle: {middle_text}"

    # Use the Phi4 model to extract information
    prompt = (
        f"Extract the following details from the provided text strictly in 2 words or less, translated to english. Do not include any word that is not english, just the english translation. "
        f"If you find birthplace, enter as Country. If there is additional information, enter as Town/Country, Region/Country, Provice/Country, etc. Any information is missing, return '-'.\nContext: {context}\n"
        f"Details to extract: Name, Surname, Father, Mother, Spouse, "
        f"Birthplace, Nationality, Religion, Post War Occupation"
    )

    response = model.generate([prompt])
    response_text = response.generations[0][0].text  # Extract the generated text

    # Parse the response and extract the details
    details = response_text.strip().split("\n")
    parsed_details = [detail.split(": ")[-1].strip() for detail in details]

    # Ensure all columns are filled, replace any missing or malformed data with '-'
    if len(parsed_details) < 9:
        parsed_details += ["-"] * (9 - len(parsed_details))

    # Write the data to the output Excel file
    output_sheet.append([id_value] + parsed_details)

    # Save the output workbook after processing each row
    output_wb.save(output_file)

print(f"Data has been extracted and saved to {output_file}.")